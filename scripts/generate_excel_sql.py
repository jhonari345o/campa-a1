"""Convierte los dos libros fuente de Ad Mavericks en SQL idempotente para Supabase.

Uso:
  python scripts/generate_excel_sql.py

También acepta --campaign-book, --radio-book y --output-dir. Los vacíos reales
del Excel se convierten en NULL; las tablas raw conservan una matriz JSONB por
fila y las tablas normalizadas se generan por separado.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import load_workbook


DEFAULT_CAMPAIGN = Path(r"C:\Users\jhonm\Downloads\20072026164449 (2) (1).xlsx")
DEFAULT_RADIO = Path(r"C:\Users\jhonm\Downloads\Rankings Radios.xlsx")
UUID_NAMESPACE = uuid.UUID("0f52dcb8-99b7-4e33-b168-f1f641b808d7")
BATCH_SIZE = 250


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def workbook_id(file_sha: str) -> uuid.UUID:
    return uuid.uuid5(UUID_NAMESPACE, f"ad-mavericks-source:{file_sha}")


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def clean_cell(value: Any) -> Any:
    if is_blank(value):
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, float):
        if not math.isfinite(value):
            return None
        return value
    return value


def sql_text(value: str) -> str:
    return "'" + value.replace("'", "''") + "'"


def sql_value(value: Any) -> str:
    value = clean_cell(value)
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (int, Decimal)):
        return str(value)
    if isinstance(value, float):
        return format(value, ".15g")
    if isinstance(value, (datetime, date)):
        return sql_text(value.isoformat())
    return sql_text(str(value))


def jsonb_value(values: Sequence[Any]) -> str:
    payload = json.dumps([clean_cell(value) for value in values], ensure_ascii=False, separators=(",", ":"))
    return f"{sql_text(payload)}::jsonb"


def numeric(value: Any) -> Decimal | None:
    value = clean_cell(value)
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        return Decimal(str(value))
    try:
        return Decimal(str(value).strip().replace(",", ""))
    except Exception:
        return None


def integer(value: Any) -> int:
    number = numeric(value)
    return int(number or 0)


def chunks(items: Sequence[str], size: int = BATCH_SIZE) -> Iterable[Sequence[str]]:
    for start in range(0, len(items), size):
        yield items[start : start + size]


def insert_batches(table: str, columns: Sequence[str], rows: Sequence[str], conflict: str) -> str:
    statements: list[str] = []
    column_sql = ", ".join(columns)
    for batch in chunks(rows):
        statements.append(
            f"insert into {table} ({column_sql}) values\n"
            + ",\n".join(batch)
            + f"\n{conflict};"
        )
    return "\n\n".join(statements)


def workbook_insert(
    wb_id: uuid.UUID,
    path: Path,
    file_sha: str,
    source_label: str,
    period_start: str | None,
    period_end: str | None,
    metadata: dict[str, Any],
) -> str:
    metadata_sql = sql_text(json.dumps(metadata, ensure_ascii=False, separators=(",", ":"))) + "::jsonb"
    return f"""update public.source_workbooks
set is_current = false
where source_label = {sql_text(source_label)}
  and id <> {sql_text(str(wb_id))}::uuid;

insert into public.source_workbooks
  (id, file_name, sha256, source_label, is_current, period_start, period_end, metadata)
values
  ({sql_text(str(wb_id))}::uuid, {sql_text(path.name)}, {sql_text(file_sha)}, {sql_text(source_label)}, true,
   {sql_value(period_start)}::date, {sql_value(period_end)}::date, {metadata_sql})
on conflict (id) do update set
  file_name = excluded.file_name,
  sha256 = excluded.sha256,
  source_label = excluded.source_label,
  is_current = true,
  period_start = excluded.period_start,
  period_end = excluded.period_end,
  metadata = excluded.metadata,
  imported_at = now();"""


def raw_sheet_rows(workbook, wb_id: uuid.UUID) -> tuple[list[str], list[dict[str, Any]]]:
    rows: list[str] = []
    sheet_meta: list[dict[str, Any]] = []
    for sheet in workbook.worksheets:
        sheet_meta.append({"name": sheet.title, "rows": sheet.max_row, "columns": sheet.max_column})
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            rows.append(
                f"({sql_text(str(wb_id))}::uuid, {sql_text(sheet.title)}, {row_index}, {jsonb_value(row)})"
            )
    return rows, sheet_meta


def generate_campaign_raw(path: Path, output: Path) -> dict[str, Any]:
    file_sha = sha256(path)
    wb_id = workbook_id(file_sha)
    workbook = load_workbook(path, data_only=True, read_only=True)
    raw_rows, sheets = raw_sheet_rows(workbook, wb_id)
    source_sql = workbook_insert(
        wb_id,
        path,
        file_sha,
        "Inversión publicitaria Ecuador",
        "2026-01-01",
        "2026-06-30",
        {"source": "Excel entregado por el usuario", "sheets": sheets, "currency": "USD"},
    )
    rows_sql = insert_batches(
        "public.source_sheet_rows",
        ("workbook_id", "sheet_name", "source_row", "cells"),
        raw_rows,
        "on conflict (workbook_id, sheet_name, source_row) do update set cells = excluded.cells",
    )
    output.write_text(
        "-- Ad Mavericks One · Copia cruda del libro de inversión\n"
        "-- Ejecutar después de 00_raw_schema.sql.\n\nbegin;\n\n"
        + source_sql
        + "\n\n"
        + rows_sql
        + "\n\ncommit;\n",
        encoding="utf-8",
    )
    return {"id": str(wb_id), "sha256": file_sha, "raw_rows": len(raw_rows), "sheets": sheets}


def generate_campaign_normalized(path: Path, output: Path, wb_id: uuid.UUID) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)

    advertiser_rows: list[str] = []
    advertiser_sheet = workbook["ANUNCIANTES"]
    for source_row, row in enumerate(advertiser_sheet.iter_rows(min_row=4, values_only=True), start=4):
        if is_blank(row[0]):
            continue
        advertiser_name = str(row[0]).strip()
        if advertiser_name.casefold() == "total general":
            continue
        values = [numeric(row[index] if index < len(row) else None) for index in range(1, 9)]
        advertiser_rows.append(
            "("
            + ", ".join(
                [sql_text(str(wb_id)) + "::uuid", str(source_row), sql_text(advertiser_name)]
                + [sql_value(value) for value in values]
            )
            + ")"
        )

    media_pairs = {
        "Internet": (4, 5),
        "Prensa": (6, 7),
        "Radio": (8, 9),
        "Revista": (10, 11),
        "Suplemento": (13, 14),
        "TV": (15, 16),
        "TVPagada": (17, 18),
    }
    media_totals = {medium: [Decimal("0"), 0] for medium in media_pairs}
    media_sheet = workbook["MEDIOS"]
    agency_rows = 0
    for row in media_sheet.iter_rows(min_row=5, values_only=True):
        label = "" if is_blank(row[0]) else str(row[0]).strip()
        if not label.lower().endswith(" total"):
            continue
        agency_rows += 1
        for medium, (investment_index, count_index) in media_pairs.items():
            media_totals[medium][0] += numeric(row[investment_index]) or Decimal("0")
            media_totals[medium][1] += integer(row[count_index])

    media_rows = [
        f"({sql_text(str(wb_id))}::uuid, {sql_text(medium)}, {sql_value(total)}, {count})"
        for medium, (total, count) in media_totals.items()
    ]

    month_pairs = {
        date(2026, 1, 1): (4, 5),
        date(2026, 2, 1): (6, 7),
        date(2026, 3, 1): (8, 9),
        date(2026, 4, 1): (10, 11),
        date(2026, 5, 1): (13, 14),
        date(2026, 6, 1): (15, 16),
    }
    month_totals = {month: [Decimal("0"), 0] for month in month_pairs}
    month_sheet = workbook["MESES"]
    for row in month_sheet.iter_rows(min_row=5, values_only=True):
        label = "" if is_blank(row[0]) else str(row[0]).strip()
        if not label.lower().endswith(" total"):
            continue
        for month, (investment_index, count_index) in month_pairs.items():
            month_totals[month][0] += numeric(row[investment_index]) or Decimal("0")
            month_totals[month][1] += integer(row[count_index])

    month_rows = [
        f"({sql_text(str(wb_id))}::uuid, {sql_text(month.isoformat())}::date, {sql_value(total)}, {count})"
        for month, (total, count) in month_totals.items()
    ]

    advertiser_sql = insert_batches(
        "public.advertiser_media_investment",
        (
            "workbook_id", "source_row", "advertiser_name", "total_general", "internet", "press",
            "radio", "magazine", "supplement", "television", "pay_tv",
        ),
        advertiser_rows,
        "on conflict (workbook_id, source_row) do update set advertiser_name = excluded.advertiser_name, "
        "total_general = excluded.total_general, internet = excluded.internet, press = excluded.press, "
        "radio = excluded.radio, magazine = excluded.magazine, supplement = excluded.supplement, "
        "television = excluded.television, pay_tv = excluded.pay_tv",
    )
    media_sql = insert_batches(
        "public.market_media_summary",
        ("workbook_id", "medium", "estimated_investment", "ad_count"),
        media_rows,
        "on conflict (workbook_id, medium) do update set estimated_investment = excluded.estimated_investment, "
        "ad_count = excluded.ad_count",
    )
    month_sql = insert_batches(
        "public.market_monthly_summary",
        ("workbook_id", "period_month", "estimated_investment", "ad_count"),
        month_rows,
        "on conflict (workbook_id, period_month) do update set estimated_investment = excluded.estimated_investment, "
        "ad_count = excluded.ad_count",
    )
    output.write_text(
        "-- Ad Mavericks One · Datos normalizados del libro de inversión\n"
        "-- Ejecutar después de 01_campaign_raw.sql.\n\nbegin;\n\n"
        + f"delete from public.advertiser_media_investment where workbook_id = {sql_text(str(wb_id))}::uuid;\n"
        + f"delete from public.market_media_summary where workbook_id = {sql_text(str(wb_id))}::uuid;\n"
        + f"delete from public.market_monthly_summary where workbook_id = {sql_text(str(wb_id))}::uuid;\n\n"
        + advertiser_sql
        + "\n\n"
        + media_sql
        + "\n\n"
        + month_sql
        + "\n\ncommit;\n",
        encoding="utf-8",
    )
    return {
        "advertisers": len(advertiser_rows),
        "agency_rows_aggregated": agency_rows,
        "media_summaries": len(media_rows),
        "monthly_summaries": len(month_rows),
    }


def generate_radio(path: Path, output: Path) -> dict[str, Any]:
    file_sha = sha256(path)
    wb_id = workbook_id(file_sha)
    workbook = load_workbook(path, data_only=True, read_only=True)
    raw_rows, sheets = raw_sheet_rows(workbook, wb_id)
    source_sql = workbook_insert(
        wb_id,
        path,
        file_sha,
        "Ranking de radios Interviú",
        None,
        None,
        {"source": "Excel entregado por el usuario", "sheets": sheets, "metric_note": "Ranking contextual; no equivale automáticamente a reach de campaña."},
    )
    raw_sql = insert_batches(
        "public.source_sheet_rows",
        ("workbook_id", "sheet_name", "source_row", "cells"),
        raw_rows,
        "on conflict (workbook_id, sheet_name, source_row) do update set cells = excluded.cells",
    )

    stations: dict[str, dict[str, Any]] = {}
    audience_sheet = workbook["Radios x Audiencia"]
    for rank, row in enumerate(audience_sheet.iter_rows(min_row=4, values_only=True), start=1):
        if is_blank(row[1]):
            continue
        name = str(row[1]).strip()
        stations[name] = {
            "station": name,
            "genre": clean_cell(row[2]),
            "dial": clean_cell(row[3]),
            "frequency": clean_cell(row[4]),
            "rating": numeric(row[5]),
            "share": numeric(row[6]),
            "rating_audience": numeric(row[7]),
            "audience_rank": rank,
        }

    reach_sheet = workbook["Radios x Alcance"]
    for rank, row in enumerate(reach_sheet.iter_rows(min_row=4, values_only=True), start=1):
        if is_blank(row[1]):
            continue
        name = str(row[1]).strip()
        station = stations.setdefault(
            name,
            {
                "station": name,
                "genre": clean_cell(row[2]),
                "dial": clean_cell(row[3]),
                "frequency": clean_cell(row[4]),
                "rating": None,
                "share": None,
                "rating_audience": None,
                "audience_rank": None,
            },
        )
        station.update(
            {
                "reach_audience": numeric(row[5]),
                "reach_pct": numeric(row[6]),
                "exclusive_reach_pct": numeric(row[7]),
                "reach_rank": rank,
            }
        )

    station_rows: list[str] = []
    for station in stations.values():
        station_rows.append(
            "("
            + ", ".join(
                [
                    sql_text(str(wb_id)) + "::uuid",
                    sql_value(station["station"]),
                    sql_value(station.get("genre")),
                    sql_value(station.get("dial")),
                    sql_value(station.get("frequency")),
                    sql_value(station.get("rating")),
                    sql_value(station.get("share")),
                    sql_value(station.get("rating_audience")),
                    sql_value(station.get("reach_audience")),
                    sql_value(station.get("reach_pct")),
                    sql_value(station.get("exclusive_reach_pct")),
                    sql_value(station.get("audience_rank")),
                    sql_value(station.get("reach_rank")),
                ]
            )
            + ")"
        )
    station_sql = insert_batches(
        "public.radio_station_metrics",
        (
            "workbook_id", "station_name", "genre", "dial", "frequency", "rating", "share",
            "rating_audience", "reach_audience", "reach_pct", "exclusive_reach_pct", "audience_rank", "reach_rank",
        ),
        station_rows,
        "on conflict (workbook_id, station_name) do update set genre = excluded.genre, dial = excluded.dial, "
        "frequency = excluded.frequency, rating = excluded.rating, share = excluded.share, "
        "rating_audience = excluded.rating_audience, reach_audience = excluded.reach_audience, "
        "reach_pct = excluded.reach_pct, exclusive_reach_pct = excluded.exclusive_reach_pct, "
        "audience_rank = excluded.audience_rank, reach_rank = excluded.reach_rank",
    )
    output.write_text(
        "-- Ad Mavericks One · Ranking radial crudo y normalizado\n"
        "-- Ejecutar después de 00_raw_schema.sql; no convierte ranking en reach de campaña.\n\nbegin;\n\n"
        + source_sql
        + "\n\n"
        + raw_sql
        + "\n\n"
        + f"delete from public.radio_station_metrics where workbook_id = {sql_text(str(wb_id))}::uuid;\n\n"
        + station_sql
        + "\n\ncommit;\n",
        encoding="utf-8",
    )
    return {"id": str(wb_id), "sha256": file_sha, "raw_rows": len(raw_rows), "stations": len(station_rows), "sheets": sheets}


def validation_sql(campaign_id: str, radio_id: str, counts: dict[str, int]) -> str:
    return f"""-- Ad Mavericks One · Verificación posterior a la importación
-- Ejecutar al final. Una diferencia detiene la transacción con un mensaje claro.

do $$
begin
  if (select count(*) from public.source_sheet_rows where workbook_id = {sql_text(campaign_id)}::uuid) <> {counts['campaign_raw']} then
    raise exception 'Filas crudas de inversión incompletas';
  end if;
  if (select count(*) from public.advertiser_media_investment where workbook_id = {sql_text(campaign_id)}::uuid) <> {counts['advertisers']} then
    raise exception 'Anunciantes normalizados incompletos';
  end if;
  if (select count(*) from public.market_media_summary where workbook_id = {sql_text(campaign_id)}::uuid) <> 7 then
    raise exception 'Resumen por medios incompleto';
  end if;
  if (select count(*) from public.market_monthly_summary where workbook_id = {sql_text(campaign_id)}::uuid) <> 6 then
    raise exception 'Resumen mensual incompleto';
  end if;
  if (select count(*) from public.source_sheet_rows where workbook_id = {sql_text(radio_id)}::uuid) <> {counts['radio_raw']} then
    raise exception 'Filas crudas de radio incompletas';
  end if;
  if (select count(*) from public.radio_station_metrics where workbook_id = {sql_text(radio_id)}::uuid) <> {counts['stations']} then
    raise exception 'Radios normalizadas incompletas';
  end if;
end $$;

select file_name, sha256, imported_at, metadata
from public.source_workbooks
order by imported_at desc;
"""


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--campaign-book", type=Path, default=DEFAULT_CAMPAIGN)
    parser.add_argument("--radio-book", type=Path, default=DEFAULT_RADIO)
    parser.add_argument("--output-dir", type=Path, default=Path("supabase/imports/excel_2026"))
    args = parser.parse_args()

    if not args.campaign_book.exists():
        raise SystemExit(f"No existe el libro de inversión: {args.campaign_book}")
    if not args.radio_book.exists():
        raise SystemExit(f"No existe el libro de radios: {args.radio_book}")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    campaign = generate_campaign_raw(args.campaign_book, args.output_dir / "01_campaign_raw.sql")
    normalized = generate_campaign_normalized(
        args.campaign_book,
        args.output_dir / "02_campaign_normalized.sql",
        uuid.UUID(campaign["id"]),
    )
    radio = generate_radio(args.radio_book, args.output_dir / "03_radio.sql")
    counts = {
        "campaign_raw": campaign["raw_rows"],
        "advertisers": normalized["advertisers"],
        "radio_raw": radio["raw_rows"],
        "stations": radio["stations"],
    }
    (args.output_dir / "99_validate_import.sql").write_text(
        validation_sql(campaign["id"], radio["id"], counts),
        encoding="utf-8",
    )
    manifest = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "campaign": {**campaign, **normalized},
        "radio": radio,
        "execution_order": [
            "00_raw_schema.sql",
            "01_campaign_raw.sql",
            "02_campaign_normalized.sql",
            "03_radio.sql",
            "99_validate_import.sql",
        ],
    }
    (args.output_dir / "IMPORT_MANIFEST.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(manifest, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
